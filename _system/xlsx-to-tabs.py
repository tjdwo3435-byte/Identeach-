# -*- coding: utf-8 -*-
"""
xlsx-to-tabs.py — 엑셀(또는 구글시트에서 받은 xlsx)의 탭들을 탭별 CSV 로 푼다.

sync 로 받아온 파일과 로컬에서 만든 파일 모두 이걸 거쳐 같은 모양이 된다.
그 다음은 form-to-data.js 가 이어받는다.

    python _system/xlsx-to-tabs.py [엑셀경로]
"""
import csv
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
DEFAULT_SRC = BASE / "data" / "아이덴티치_프로그램_입력시트.xlsx"
OUT_DIR = BASE / "data" / "tabs"


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.exists():
        sys.exit(f"파일이 없습니다: {src}")

    wb = load_workbook(src, data_only=True)

    # 이전 결과를 지우고 새로 쓴다 (지워진 탭이 남아있으면 안 되니까)
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    OUT_DIR.mkdir(parents=True)

    order = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            while cells and cells[-1] == "":
                cells.pop()
            rows.append(cells)
        while rows and not rows[-1]:
            rows.pop()

        # 탭 이름을 그대로 파일명으로 쓴다 (윈도우에서 못 쓰는 글자만 바꿈)
        safe = "".join("_" if c in '<>:"/\\|?*' else c for c in ws.title)
        with (OUT_DIR / f"{safe}.csv").open("w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)

        order.append(safe)
        mark = "(안내)" if ws.title.startswith("_") else ""
        print(f"  {ws.title:<18} {len(rows):>3}행 {mark}")

    # 탭이 놓인 순서 = 홈페이지 목록에 뜨는 순서. 파일 이름만으로는 알 수 없어 따로 남긴다.
    (OUT_DIR / "_탭순서.json").write_text(
        json.dumps(order, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"\n탭 {len(wb.worksheets)}개 → {OUT_DIR}")


if __name__ == "__main__":
    main()
