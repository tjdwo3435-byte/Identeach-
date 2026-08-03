# -*- coding: utf-8 -*-
"""
make-form.py — 프로그램 입력용 엑셀을 만든다. (탭 하나 = 프로그램 하나)

만들어지는 탭
    _읽어보기        이 시트가 뭔지, 새 프로그램 추가하는 법, 어느 칸이 어디에 나오는지
    _예시            다 채워진 예시. 이걸 보고 따라 쓰면 된다.
    _양식            빈 양식. 새 프로그램은 이 탭을 복사해서 쓴다.
    (프로그램명)…    지금 홈페이지에 있는 프로그램들

밑줄(_)로 시작하는 탭은 빌드가 무시한다.

    python _system/make-form.py
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
SCHEMA = json.loads((BASE / "form-schema.json").read_text(encoding="utf-8"))
PROGRAMS = json.loads((BASE / "data" / "programs.json").read_text(encoding="utf-8"))
OUT = BASE / "data" / "아이덴티치_프로그램_입력시트.xlsx"

# 색
NAVY = "1F2392"
YELLOW = "FFD84D"
WARM = "F7F7FA"
GRAY = "6F7287"
LINE = "D8D8E4"

F_TITLE = Font(size=16, bold=True, color=NAVY)
F_SECTION = Font(size=11, bold=True, color="FFFFFF")
F_LABEL = Font(size=10, bold=True, color="14142B")
F_HINT = Font(size=9, color=GRAY)
F_DESC = Font(size=9, color=GRAY, italic=True)
F_HEAD = Font(size=9, bold=True, color=NAVY)

FILL_SECTION = PatternFill("solid", fgColor=NAVY)
FILL_LABEL = PatternFill("solid", fgColor=WARM)
FILL_INPUT = PatternFill("solid", fgColor="FFFFFF")
FILL_HEAD = PatternFill("solid", fgColor="EDEDF5")

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)

# 값이 들어가는 열은 B~F, 도움말은 G
VAL_FIRST, VAL_LAST, HINT_COL = 2, 6, 7
WIDTHS = {"A": 15, "B": 11, "C": 19, "D": 40, "E": 23, "F": 14, "G": 46}


# ── 시트 공통 ────────────────────────────────────────────
def setup(ws):
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False


def put_title(ws, row, text, sub=""):
    ws.cell(row=row, column=1, value=text).font = F_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=HINT_COL)
    ws.row_dimensions[row].height = 26
    row += 1
    if sub:
        ws.cell(row=row, column=1, value=sub).font = F_DESC
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=HINT_COL)
        row += 1
    return row + 1


def put_section(ws, row, title, desc):
    c = ws.cell(row=row, column=1, value=f"  {title}")
    c.font, c.fill = F_SECTION, FILL_SECTION
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=HINT_COL)
    ws.row_dimensions[row].height = 20
    row += 1
    if desc:
        c = ws.cell(row=row, column=1, value=f"  {desc}")
        c.font, c.alignment = F_DESC, WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=HINT_COL)
        row += 1
    return row


def put_field(ws, row, label, value, hint):
    """A=항목이름, B~F=값(하나로 합침), G=도움말"""
    a = ws.cell(row=row, column=1, value=label)
    a.font, a.fill, a.border, a.alignment = F_LABEL, FILL_LABEL, BOX, WRAP_TOP

    v = ws.cell(row=row, column=VAL_FIRST, value=value or None)
    v.font, v.fill, v.alignment = Font(size=10), FILL_INPUT, WRAP_TOP
    ws.merge_cells(start_row=row, start_column=VAL_FIRST, end_row=row, end_column=VAL_LAST)
    for c in range(VAL_FIRST, VAL_LAST + 1):
        ws.cell(row=row, column=c).border = BOX

    h = ws.cell(row=row, column=HINT_COL, value=hint)
    h.font, h.alignment = F_HINT, WRAP_TOP

    # 줄이 여러 개면 행 높이를 늘려 다 보이게 한다
    lines = str(value or "").count("\n") + 1
    ws.row_dimensions[row].height = max(20, min(96, 15 * lines + 5))
    return row + 1


def put_table(ws, row, key, label, columns, nrows, values):
    """표: A열에 항목 이름, B열부터 각 칸"""
    if label:
        c = ws.cell(row=row, column=1, value=label)
        c.font, c.alignment = F_HEAD, WRAP_TOP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=HINT_COL)
        row += 1

    # 머리글
    a = ws.cell(row=row, column=1, value=key)
    a.font, a.fill, a.border, a.alignment = F_LABEL, FILL_LABEL, BOX, WRAP_TOP
    for i, col in enumerate(columns):
        c = ws.cell(row=row, column=VAL_FIRST + i, value=col["label"])
        c.font, c.fill, c.border, c.alignment = F_HEAD, FILL_HEAD, BOX, CENTER
    hint = " / ".join(f'{c["label"]}: {c["hint"]}' for c in columns)
    h = ws.cell(row=row, column=HINT_COL, value=hint)
    h.font, h.alignment = F_HINT, WRAP_TOP
    ws.row_dimensions[row].height = max(20, 13 * (len(hint) // 40 + 1))
    row += 1

    # 값 (모자라면 빈 줄로 채움 — 나중에 추가하기 쉽게)
    for r in range(nrows):
        vals = values[r] if r < len(values) else [""] * len(columns)
        tallest = 1
        for i in range(len(columns)):
            v = vals[i] if i < len(vals) else ""
            c = ws.cell(row=row, column=VAL_FIRST + i, value=v or None)
            c.font, c.fill, c.border, c.alignment = Font(size=10), FILL_INPUT, BOX, WRAP_TOP
            tallest = max(tallest, str(v or "").count("\n") + 1)
        ws.cell(row=row, column=1).border = BOX
        ws.row_dimensions[row].height = max(18, min(80, 15 * tallest + 4))
        row += 1

    return row + 1


# ── programs.json → 폼 값 ────────────────────────────────
LEGACY = {"career/talkshow", "career/job-lab", "career/ceo-talk"}


def info_val(info, label):
    for x in info or []:
        if x.get("label") == label:
            return x.get("value", "")
    return ""


def photo_folder(d):
    """모든 사진이 같은 폴더에 있으면 그 폴더를 돌려준다."""
    srcs = []
    if d.get("activity", {}).get("image", {}).get("src"):
        srcs.append(d["activity"]["image"]["src"])
    srcs += [i["src"] for i in d.get("heroImages", []) if i.get("src")]
    srcs += [i["src"] for i in d.get("gallery", {}).get("images", []) if i.get("src")]
    dirs = {s.rsplit("/", 1)[0] for s in srcs if "/" in s}
    return dirs.pop() if len(dirs) == 1 else ""


def rel(src, folder):
    if folder and src.startswith(folder + "/"):
        return src[len(folder) + 1:]
    return src


def to_form(slug, d):
    folder = photo_folder(d)
    card = d.get("_card", {})
    ss = d.get("structuresSection") or {}

    fields = {
        "사용": "" if slug in LEGACY else "O",
        "ID": slug,
        "분류": d.get("category") or (d.get("crumbs", [{}, {}])[1].get("label", "")),
        "프로그램명": d.get("title", ""),
        "한줄배지": d.get("tagline", ""),
        "소개문": d.get("description", ""),
        "대상": info_val(d.get("info"), "대상"),
        "시간": info_val(d.get("info"), "시간"),
        "인원": info_val(d.get("info"), "인원"),
        "장소": info_val(d.get("info"), "장소"),
        "개요제목": d.get("activity", {}).get("headline", ""),
        "개요내용": "\n".join(d.get("activity", {}).get("bullets", [])),
        "활동유형제목": ss.get("title", ""),
        "사진폴더": folder,
        "개요사진": rel(d.get("activity", {}).get("image", {}).get("src", ""), folder),
        "개요사진설명": d.get("activity", {}).get("image", {}).get("alt", ""),
        "카드썸네일": card.get("카드썸네일", ""),
        "카드배지": card.get("카드배지", ""),
        "카드배지색": card.get("카드배지색", ""),
        "카드제목": card.get("카드제목", ""),
        "카드설명": card.get("카드설명", ""),
        "카드태그": card.get("카드태그", ""),
        "준비중": card.get("준비중", ""),
        "검색제목": d.get("pageTitle", ""),
        "검색설명": d.get("pageDesc", ""),
        # 제안서·블로그용은 새로 생긴 항목이라 비어 있다 (사람이 채운다)
        "교육목표": "", "기대효과": "", "준비물": "", "진행인력": "",
        "단가": "", "안전": "", "블로그키워드": "", "성수기": "",
        "비고": "구형 구조 — 표준 전환 필요" if slug in LEGACY else "",
    }

    tables = {
        "활용장면": [
            [c.get("icon", ""), "O" if c.get("rec") else "", c.get("title", ""), c.get("desc", "")]
            for c in d.get("useCases", [])
        ],
        "진행흐름": [
            [s.get("icon", ""), s.get("proc", {}).get("title", ""), s.get("proc", {}).get("desc", ""),
             s.get("xp", {}).get("action", ""), s.get("xp", {}).get("detail", "")]
            for s in d.get("flowSteps", [])
        ],
        "활동유형": [[it.get("name", ""), it.get("desc", "")] for it in ss.get("items", [])],
        "대표사진": [[rel(i["src"], folder), i.get("alt", "")] for i in d.get("heroImages", [])],
        "갤러리사진": [
            [rel(i["src"], folder), i.get("alt", ""), i.get("caption", "")]
            for i in d.get("gallery", {}).get("images", [])
        ],
    }
    return fields, tables


# ── 탭 만들기 ────────────────────────────────────────────
def build_tab(wb, name, title, sub, fields, tables, use_example=False):
    ws = wb.create_sheet(title=name)
    setup(ws)
    row = put_title(ws, 1, title, sub)

    for sec in SCHEMA["sections"]:
        row = put_section(ws, row, sec["title"], sec.get("desc", ""))

        for f in sec.get("fields", []):
            val = f.get("example", "") if use_example else fields.get(f["key"], "")
            row = put_field(ws, row, f["label"], val, f.get("hint", ""))

        for t in ([sec["table"]] if "table" in sec else []) + sec.get("tables", []):
            vals = t.get("example", []) if use_example else tables.get(t["key"], [])
            row = put_table(ws, row, t["key"], t.get("label", ""), t["columns"], t["rows"], vals)

        row += 1

    return ws


def build_readme(wb):
    ws = wb.create_sheet(title="_읽어보기")
    setup(ws)
    row = put_title(
        ws, 1,
        "프로그램 입력 시트",
        "탭 하나 = 프로그램 하나입니다. 여기에 적은 내용이 홈페이지·블로그·제안서에 쓰입니다.",
    )

    row = put_section(ws, row, "새 프로그램을 만들려면", "")
    for i, step in enumerate([
        "아래 [_양식] 탭에서 마우스 오른쪽 → '복사본 만들기' 를 누릅니다.",
        "새로 생긴 탭의 이름을 프로그램 이름으로 바꿉니다. (예: 도미노 챌린지)",
        "[_예시] 탭을 옆에 띄워두고, 같은 자리에 우리 프로그램 내용을 채웁니다.",
        "맨 위 '사용' 칸에 O 를 적으면 홈페이지에 올라갑니다. 아직이면 비워두세요.",
        "다 채웠으면 담당자에게 알려주세요. 홈페이지 반영은 명령 두 줄이면 끝납니다.",
    ], start=1):
        c = ws.cell(row=row, column=1, value=f"{i}단계")
        c.font, c.fill, c.border, c.alignment = F_LABEL, FILL_LABEL, BOX, WRAP_TOP
        v = ws.cell(row=row, column=VAL_FIRST, value=step)
        v.font, v.alignment = Font(size=10), WRAP_TOP
        ws.merge_cells(start_row=row, start_column=VAL_FIRST, end_row=row, end_column=HINT_COL)
        row += 1
    row += 1

    row = put_section(ws, row, "채울 때 꼭 지킬 것", "")
    for rule in [
        "밑줄(_)로 시작하는 탭은 건드리지 마세요. 안내용이라 홈페이지에 안 올라갑니다.",
        "'주소(ID)' 는 프로그램마다 달라야 합니다. 겹치면 먼저 것이 덮어써집니다.",
        "칸 안에서 줄을 바꾸려면 Alt+Enter 를 쓰세요.",
        "칸 맨 앞에 작은따옴표(')를 쓰려면 두 번('') 치세요. 한 번만 치면 구글 시트가 먹어버립니다.",
        "표에서 빈 줄은 그냥 두면 됩니다. 알아서 건너뜁니다.",
        "사진은 파일 이름만 적습니다. 폴더는 위쪽 '사진폴더' 칸에 한 번만 적으면 됩니다.",
    ]:
        c = ws.cell(row=row, column=1, value="•")
        c.font, c.alignment = F_LABEL, CENTER
        v = ws.cell(row=row, column=VAL_FIRST, value=rule)
        v.font, v.alignment = Font(size=10), WRAP_TOP
        ws.merge_cells(start_row=row, start_column=VAL_FIRST, end_row=row, end_column=HINT_COL)
        row += 1
    row += 1

    row = put_section(ws, row, "어느 칸이 어디에 나오나요", "")
    head = ["항목", "쓰이는 곳", "설명"]
    for i, h in enumerate(head):
        c = ws.cell(row=row, column=1 + (0 if i == 0 else (1 if i == 1 else 2)), value=h)
        c.font, c.fill, c.border, c.alignment = F_HEAD, FILL_HEAD, BOX, CENTER
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=HINT_COL)
    row += 1

    USE_TEXT = {
        "홈": "홈페이지 상세",
        "목록": "목록 카드",
        "검색": "검색·공유",
        "내부": "제안서·블로그",
    }
    for sec in SCHEMA["sections"]:
        for f in sec.get("fields", []):
            ws.cell(row=row, column=1, value=f["label"]).font = Font(size=9, bold=True)
            u = ws.cell(row=row, column=2, value=USE_TEXT.get(f.get("use", ""), ""))
            u.font = Font(size=9, color=NAVY if f.get("use") != "내부" else GRAY)
            d = ws.cell(row=row, column=3, value=f.get("hint", ""))
            d.font, d.alignment = F_HINT, WRAP_TOP
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=HINT_COL)
            for c in range(1, HINT_COL + 1):
                ws.cell(row=row, column=c).border = BOX
            row += 1
    return ws


def safe_tab_name(name, used):
    """엑셀 탭 이름 규칙: 31자 이하, : \\ / ? * [ ] 못 씀"""
    n = re.sub(r"[:\\/?*\[\]]", " ", name).strip()[:31] or "프로그램"
    base, i = n, 2
    while n in used:
        n = f"{base[:28]} {i}"
        i += 1
    used.add(n)
    return n


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb)

    # 예시 탭 — 스키마의 example 값으로 채운다
    build_tab(wb, "_예시", "예시 — 미니 스포츠데이",
              "이렇게 채우면 됩니다. 오른쪽 회색 글씨가 각 칸 설명입니다.",
              {}, {}, use_example=True)

    # 빈 양식 탭
    build_tab(wb, "_양식", "새 프로그램",
              "이 탭을 복사해서 쓰세요. 탭 이름은 프로그램 이름으로 바꾸면 됩니다.",
              {}, {})

    # 프로그램 탭들 — 목록 페이지에 놓인 순서대로
    used = {"_읽어보기", "_예시", "_양식"}
    order = sorted(PROGRAMS.items(), key=lambda kv: kv[1].get("_order", 999))
    for slug, d in order:
        fields, tables = to_form(slug, d)
        name = safe_tab_name(fields["프로그램명"] or slug, used)
        build_tab(wb, name, fields["프로그램명"], f"주소: identeach.co.kr/{slug}", fields, tables)
        print(f"  탭  {name:<16} ({slug})")

    wb.save(OUT)
    print(f"\n탭 {len(wb.sheetnames)}개 → {OUT.name}  ({OUT.stat().st_size:,} bytes)")
    print(f"   {OUT}")


if __name__ == "__main__":
    main()
